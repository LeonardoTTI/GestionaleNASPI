from openpyxl import load_workbook 
import datetime
import csv
import sys

def main():
	#pattern_sporco=["N", "Centro", "Codice Fiscale", "Cognome Nome", "Pratica n. pratica", "Importo", "Sindacato", "Anno"]
	pattern_pulito=["N", "Codice Fiscale", "Cognome Nome", "Pratica n. pratica", "Importo"]
	conta_pulito = 0
	book = load_workbook(sys.argv[1])
	sheet = book.worksheets[0] #per file con più di uno sheet
	#sheet = book.active per file con un solo sheet
	output_path = "./saves/"+str(datetime.datetime.today()).split(" ")[0]+".csv"
	#print(output_path)




	for i in range(0,5):
		if sheet.cell(row = 1, column = i+1).value==pattern_pulito[i]:
			#print(str(sheet.cell(row = 1, column = i+1).value))
			conta_pulito+=1
			
	#print(conta_pulito)
	last_proc = 0
	if conta_pulito==5:
		f = open(output_path, "w", encoding='UTF-8', newline='')
		writer_csv = csv.writer(f)
		tmp = []
		for t_row in range(1,sheet.max_row+1):
			if not (last_proc+1==int(sheet.cell(row = t_row, column = 1).value)):
				tmp.append(str(last_proc+1==int(sheet.cell(row = t_row, column = 1).value)))
				for col in range(2,6):
					tmp.append("")
				writer_csv.writerow(tmp)
			else:
				for col in range(1,6):
					tmp.append(str(sheet.cell(row = t_row, column = col).value))
				if not (str(sheet.cell(row = t_row, column = 4).value)=='None' or str(sheet.cell(row = t_row, column = 1).value)=='N'):
					writer_csv.writerow(tmp)
					#print(tmp)
			tmp = []
			last_proc = int(sheet.cell(row = t_row, column = 1).value)
			print(last_proc)
		f.close()	
	else:
		f = open(output_path, "w", encoding='UTF-8', newline='')
		writer_csv = csv.writer(f)
		tmp = []
		for t_row in range(1,sheet.max_row+1):
			if not (str(sheet.cell(row = t_row, column = 1).value)=='N' or str(sheet.cell(row = t_row, column = 4).value)=='None'):
				if not (last_proc+1==int(sheet.cell(row = t_row, column = 1).value)):
					tmp.append(str(last_proc+1==int(sheet.cell(row = t_row, column = 1).value)))
					for col in range(3,7):
						tmp.append(str(sheet.cell(row = t_row, column = col).value))
					writer_csv.writerow(tmp)
				else:
					for col in range(1,7):
						if not(col==2):
							tmp.append(str(sheet.cell(row = t_row, column = col).value))
					if not (str(sheet.cell(row = t_row, column = 4).value)=='None' or str(sheet.cell(row = t_row, column = 1).value)=='N'):
						writer_csv.writerow(tmp)
						#print(tmp)
				last_proc = int(sheet.cell(row = t_row, column = 1).value)
			tmp = []
			print(last_proc)
		f.close()	


	book.close()

	print("Finito con successo!")
	
	return 0

if len(sys.argv) < 2:
	sys.exit("Bad usage: inserire come argomento il nome del filde da convertire in csv")
main()	

